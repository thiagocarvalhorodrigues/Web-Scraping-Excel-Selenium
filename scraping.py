from selenium import webdriver
import os
import openpyxl
import time


class BuscarPrecos:
    def __init__(self):
        self.driver = webdriver.Chrome(executable_path=os.getcwd() + os.sep + 'chromedriver.exe')

    def Iniciar(self):
        self.produto = input('Qual produto buscar?')
        #self.driver.get('https://sc.olx.com.br/florianopolis-e-regiao/grande-florianopolis?q=memoria%20ddr3%208gb')
        self.criar_planilha()
        self.encontrar_valores_na_pagina()

    def criar_planilha(self):
        self.planilha = openpyxl.Workbook()
        self.planilha.create_sheet('Valores')
        self.planilha_valores = self.planilha['Valores']
        self.planilha_valores.cell(row=1, column=1, value='Titulo')
        self.planilha_valores.cell(row=1, column=2, value='Localizacao')
        self.planilha_valores.cell(row=1, column=3, value='Precos')

    def encontrar_valores_na_pagina(self):
        try:
            for self.paginas in range(1,2):
                self.navegar_para_proxima_pagina()

                self.titulo = self.driver.find_elements_by_xpath('//h2[@class="fnmrjs-10 deEIZJ"]')
                self.localizacao = self.driver.find_elements_by_xpath('//p[@class="fnmrjs-13 hdwqVC"]')
                self.precos = self.driver.find_elements_by_xpath('//p[@class="fnmrjs-16 jqSHIm"]')
                self.armazenar_valores_em_planilha()




        except Exception as erro:
            print('Fim')


    def navegar_para_proxima_pagina(self):
        time.sleep(7)

        self.driver.get(f'https://sc.olx.com.br/florianopolis-e-regiao/grande-florianopolis?o={self.paginas}&q={self.produto}')



    def armazenar_valores_em_planilha(self):
        for indice in range(0, len(self.titulo)):
            nova_linha = [self.titulo[indice].text,
                            self.localizacao[indice].text, self.precos[indice].text]
            self.planilha_valores.append(nova_linha)
        self.planilha.save('Preços memória ddr3.xlsx')


root = BuscarPrecos()
root.Iniciar()